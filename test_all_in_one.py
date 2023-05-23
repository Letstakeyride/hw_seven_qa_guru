import os
import csv
import time
import zipfile

import xlrd
import os.path
import requests

from pypdf import PdfReader
from selene import have
from selenium import webdriver
from openpyxl import load_workbook
from selene.support.shared import browser

from os_path.os_path_scripts import tmp as RESOURCES_PATH



# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_csv():
    csv_path = os.path.join(RESOURCES_PATH, 'example.csv')
    with open(csv_path, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(csv_path) as csvfile:
        csvreader = csv.reader(csvfile)
        arr = []
        for row in csvreader:
            arr.append(row)
    assert arr[0] == ['Anna', 'Pavel', 'Peter']


def test_download_with_browser():
    browser.open("https://github.com/pytest-dev/pytest")
    browser.should(have.url("https://github.com/pytest-dev/pytest"))

    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(5)

    download_file = os.path.join(RESOURCES_PATH, 'pytest-main.zip')
    assert os.path.exists(download_file)


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    r = requests.get(url)
    path_t = os.path.join(RESOURCES_PATH, 'selenium_logo.png')
    with open('selenium_logo.png', 'wb') as file:
        file.write(r.content)

    size = os.path.getsize(path_t)

    assert size == 30803


def test_pdf():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    pdf_path = os.path.join(RESOURCES_PATH, 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(pdf_path)

    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    assert number_of_pages == 412, "The PDF should have 412 pages"
    assert 'pytest Documentation' in text, "The first page should have 'pytest Documentation' text"


def test_xls():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    xls_path = os.path.join(RESOURCES_PATH, 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(xls_path)
    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']
    sheet = book.sheet_by_index(0)
    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'
    expected_rows = [
        ['0.0', 'First Name', 'Last Name', 'Gender', 'Country', 'Age', 'Date', 'Id'],
        ['1.0', 'Dulce', 'Abril', 'Female', 'United States', '32.0', '15/10/2017', '1562.0'],
        ['2.0', 'Mara', 'Hashimoto', 'Female', 'Great Britain', '25.0', '16/08/2016', '1582.0'],
        ['3.0', 'Philip', 'Gent', 'Male', 'France', '36.0', '21/05/2015', '2587.0'],
        ['4.0', 'Kathleen', 'Hanner', 'Female', 'United States', '25.0', '15/10/2017', '3549.0'],
        ['5.0', 'Nereida', 'Magwood', 'Female', 'United States', '58.0', '16/08/2016', '2468.0'],
        ['6.0', 'Gaston', 'Brumm', 'Male', 'United States', '24.0', '21/05/2015', '2554.0'],
        ['7.0', 'Etta', 'Hurn', 'Female', 'Great Britain', '56.0', '15/10/2017', '3598.0'],
        ['8.0', 'Earlean', 'Melgar', 'Female', 'United States', '27.0', '16/08/2016', '2456.0'],
        ['9.0', 'Vincenza', 'Weiland', 'Female', 'United States', '40.0', '21/05/2015', '6548.0']
    ]
    for i, row in enumerate(sheet.get_rows()):
        expected_row = [str(cell.value) if isinstance(cell.value, float) else cell.value for cell in row]
        assert expected_row == expected_rows[i]


def test_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    xlsx_path = os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'


def test_add_files_to_zip():
    zip_name = "test.zip"
    with zipfile.ZipFile(zip_name, "w") as zip_file:
        for file_name in os.listdir(RESOURCES_PATH):
            file_path = os.path.join(RESOURCES_PATH, file_name)
            zip_file.write(file_path, file_name)

    with zipfile.ZipFile(zip_name, "r") as zip_file:
        for file_name in os.listdir(RESOURCES_PATH):
            assert file_name in zip_file.namelist()
